import sys
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

try:
    wb = openpyxl.load_workbook('Orçamento-manual de padrinho.xlsx', data_only=False)
    for sheet_name in wb.sheetnames:
        print(f"--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print(f"{cell.coordinate}: {cell.value}")
except Exception as e:
    print(f"Error: {e}")
